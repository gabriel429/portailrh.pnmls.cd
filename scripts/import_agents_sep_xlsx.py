import argparse
import json
import os
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import pymysql
from openpyxl import load_workbook


PROVINCE_DEFINITIONS = {
    'bas uele': ('BUE', 'Bas-Uele'),
    'equateur': ('EQU', 'Equateur'),
    'haut katanga': ('HAU', 'Haut-Katanga'),
    'haut lomami': ('HLO', 'Haut-Lomami'),
    'haut uele': ('HUE', 'Haut-Uele'),
    'ituri': ('ITU', 'Ituri'),
    'kasai': ('KAS', 'Kasai'),
    'kasai central': ('KCE', 'Kasai Central'),
    'kasai oriental': ('KOR', 'Kasai Oriental'),
    'kinshasa': ('KIN', 'Kinshasa'),
    'kongo central': ('KOC', 'Kongo Central'),
    'kwango': ('KWA', 'Kwango'),
    'kwilu': ('KWI', 'Kwilu'),
    'lomami': ('LOM', 'Lomami'),
    'lualaba': ('LOF', 'Lualaba'),
    'mai ndombe': ('MND', 'Mai-Ndombe'),
    'maniema': ('MAN', 'Maniema'),
    'mongala': ('MON', 'Mongala'),
    'nord kivu': ('NOR', 'Nord-Kivu'),
    'nord ubangi': ('NUB', 'Nord-Ubangi'),
    'sankuru': ('SAN', 'Sankuru'),
    'sud kivu': ('SUD', 'Sud-Kivu'),
    'sud ubangi': ('SUB', 'Sud-Ubangi'),
    'tanganyika': ('TAN', 'Tanganyika'),
    'tshopo': ('TSP', 'Tshopo'),
    'tshuapa': ('TSH', 'Tshuapa'),
}

ORGANE_DEFINITIONS = {
    'secretariat executif national': 'Secrétariat Exécutif National',
    'secretariat executif provincial': 'Secrétariat Exécutif Provincial',
    'secretariat executif local': 'Secrétariat Exécutif Local',
}

DEFAULT_BIRTH_YEAR_DELTA = 37


def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("'", ' ')
    text = text.replace('-', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.lower().strip()


def clean_text(value, allow_empty=False):
    if value is None:
        return None if not allow_empty else ''
    text = str(value).strip()
    text = re.sub(r'\s+', ' ', text)
    if text == '' or normalize_text(text) in {'vide', 'nu', 'n/a', 'none'}:
        return None if not allow_empty else ''
    return text


def stringify_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=' ')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip('0').rstrip('.')
    return str(value).strip()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_year(value):
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r'(19\d{2}|20\d{2})', text)
    return int(match.group(1)) if match else None


def valid_email(value):
    text = clean_text(value)
    if not text:
        return None
    if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', text):
        return text.lower()
    return None


def build_placeholder_email(row_number):
    return f'agent.sep.import.{row_number:04d}@pnmls.local'


def ensure_unique_email(candidate, existing_email_to_id, target_id, used_emails, row_number):
    email = candidate or build_placeholder_email(row_number)
    while True:
        owner = existing_email_to_id.get(email)
        if (owner is None or owner == target_id) and email not in used_emails:
            used_emails.add(email)
            return email, candidate is None
        email = build_placeholder_email(row_number)
        row_number += 10000


def choose_primary_email(row, idx, existing_email_to_id, target_id, used_emails, row_number):
    professional = valid_email(row[idx['Email institutionnel']])
    private = valid_email(row[idx['Email prive']])
    for candidate in (professional, private):
        if candidate:
            owner = existing_email_to_id.get(candidate)
            if owner is None or owner == target_id:
                used_emails.add(candidate)
                return candidate, False
    return ensure_unique_email(None, existing_email_to_id, target_id, used_emails, row_number)


def compute_birth_and_engagement(row, idx, anomaly_report):
    raw_birth = row[idx['Annee naissance']]
    raw_engagement = row[idx['Annee engagement']]
    birth_date = parse_date(raw_birth)
    engagement_year = parse_year(raw_engagement)

    if birth_date and birth_date.year < 1945:
        if engagement_year:
            derived_year = max(1945, engagement_year - DEFAULT_BIRTH_YEAR_DELTA)
            birth_date = date(derived_year, birth_date.month, birth_date.day)
            anomaly_report['birth_year_derived_from_placeholder'] += 1
        else:
            birth_date = date(1970, birth_date.month, birth_date.day)
            anomaly_report['birth_year_placeholder_without_engagement'] += 1

    if birth_date is None:
        if engagement_year:
            birth_date = date(max(1945, engagement_year - DEFAULT_BIRTH_YEAR_DELTA), 1, 1)
            anomaly_report['birth_date_missing_derived'] += 1
        else:
            birth_date = date(1970, 1, 1)
            anomaly_report['birth_date_missing_defaulted'] += 1

    if engagement_year is None:
        engagement_year = min(date.today().year, birth_date.year + DEFAULT_BIRTH_YEAR_DELTA)
        anomaly_report['engagement_year_missing_derived'] += 1

    return birth_date, birth_date.year, engagement_year, date(engagement_year, 1, 1)


def normalize_organe(value):
    normalized = normalize_text(value)
    return ORGANE_DEFINITIONS.get(normalized, 'Secrétariat Exécutif Provincial')


def normalize_status(value):
    normalized = normalize_text(value)
    if normalized in {'actif', 'suspendu', 'ancien'}:
        return normalized
    return 'actif'


def next_pnmls_code(row_number):
    return f'AGT-SEP-{row_number:04d}'


def ensure_provinces(cur, workbook_provinces):
    cur.execute('SELECT id, code, nom FROM provinces')
    provinces = cur.fetchall()
    province_map = {normalize_text(name): province_id for province_id, _, name in provinces}
    created = []

    for workbook_name in sorted(workbook_provinces):
        normalized = normalize_text(workbook_name)
        if normalized in province_map:
            continue
        code, canonical_name = PROVINCE_DEFINITIONS.get(normalized, (normalized[:3].upper(), workbook_name))
        description = f'Province de {canonical_name}'
        cur.execute(
            'INSERT INTO provinces (code, nom, description, created_at, updated_at) VALUES (%s, %s, %s, NOW(), NOW())',
            (code, canonical_name, description),
        )
        province_map[normalized] = cur.lastrowid
        created.append(canonical_name)

    return province_map, created


def load_grade_map(cur):
    cur.execute('SELECT id, libelle FROM grades')
    return {normalize_text(libelle): grade_id for grade_id, libelle in cur.fetchall()}


def load_existing_agents(cur):
    cur.execute('SELECT id, matricule_etat, email, nom, postnom, prenom, province_id, organe, matricule_pnmls FROM agents')
    rows = cur.fetchall()
    by_matricule = {}
    by_email = {}
    by_identity = {}

    for agent_id, matricule_etat, email, nom, postnom, prenom, province_id, organe, matricule_pnmls in rows:
        if matricule_etat:
            by_matricule[normalize_text(matricule_etat)] = agent_id
        if email:
            by_email[email.lower()] = agent_id
        identity = (
            normalize_text(nom),
            normalize_text(postnom),
            normalize_text(prenom),
            province_id or 0,
            normalize_text(organe),
        )
        by_identity[identity] = (agent_id, matricule_pnmls)

    return rows, by_matricule, by_email, by_identity


def fetch_existing_pnmls(cur, agent_id):
    cur.execute('SELECT matricule_pnmls FROM agents WHERE id = %s', (agent_id,))
    row = cur.fetchone()
    return row[0] if row else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', required=True)
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--report', default='storage/logs/agent_sep_import_report.json')
    parser.add_argument('--db-host', default=os.getenv('IMPORT_DB_HOST', '127.0.0.1'))
    parser.add_argument('--db-port', type=int, default=int(os.getenv('IMPORT_DB_PORT', '3306')))
    parser.add_argument('--db-user', default=os.getenv('IMPORT_DB_USER', 'root'))
    parser.add_argument('--db-password', default=os.getenv('IMPORT_DB_PASSWORD', ''))
    parser.add_argument('--db-name', default=os.getenv('IMPORT_DB_NAME', 'portailrh_pnmls'))
    args = parser.parse_args()

    workbook_path = Path(args.xlsx)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {header: i for i, header in enumerate(headers)}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    workbook_provinces = {clean_text(row[idx['Province']]) for row in rows if clean_text(row[idx['Province']])}

    conn = pymysql.connect(
        host=args.db_host,
        port=args.db_port,
        user=args.db_user,
        password=args.db_password,
        database=args.db_name,
        charset='utf8mb4',
        autocommit=False,
    )
    cur = conn.cursor()

    province_map, created_provinces = ensure_provinces(cur, workbook_provinces)
    grade_map = load_grade_map(cur)
    _, existing_by_matricule, existing_by_email, existing_by_identity = load_existing_agents(cur)

    anomaly_report = defaultdict(int)
    action_report = defaultdict(int)
    used_emails = set()

    for row_number, row in enumerate(rows, start=1):
        matricule_etat = clean_text(row[idx['Matricule Etat']])
        if matricule_etat and normalize_text(matricule_etat) == 'nu':
            matricule_etat = None

        province_name = clean_text(row[idx['Province']]) or 'Kinshasa'
        province_id = province_map.get(normalize_text(province_name))

        nom = clean_text(row[idx['Nom']]) or 'Non renseigne'
        postnom = clean_text(row[idx['Postnom']])
        prenom = clean_text(row[idx['Prenom']]) or 'Non renseigne'
        organe = normalize_organe(row[idx['Organe']])

        target_id = None
        if matricule_etat:
            target_id = existing_by_matricule.get(normalize_text(matricule_etat))

        if target_id is None:
            for candidate in (valid_email(row[idx['Email institutionnel']]), valid_email(row[idx['Email prive']])):
                if candidate and candidate in existing_by_email:
                    target_id = existing_by_email[candidate]
                    break

        if target_id is None:
            identity = (normalize_text(nom), normalize_text(postnom), normalize_text(prenom), province_id or 0, normalize_text(organe))
            match = existing_by_identity.get(identity)
            target_id = match[0] if match else None

        primary_email, generated_email = choose_primary_email(row, idx, existing_by_email, target_id, used_emails, row_number)
        if generated_email:
            anomaly_report['generated_primary_email'] += 1

        email_prive = valid_email(row[idx['Email prive']])
        email_professionnel = valid_email(row[idx['Email institutionnel']])
        telephone = clean_text(row[idx['Telephone']])
        fonction = clean_text(row[idx['Fonction']]) or clean_text(row[idx['Poste actuel']]) or 'Non renseigne'
        poste_actuel = clean_text(row[idx['Poste actuel']]) or fonction
        grade_etat = clean_text(row[idx['Grade Etat']])
        grade_id = grade_map.get(normalize_text(grade_etat)) if grade_etat else None
        if grade_etat and grade_id is None:
            anomaly_report['unmapped_grade'] += 1

        birth_date, birth_year, engagement_year, hire_date = compute_birth_and_engagement(row, idx, anomaly_report)

        provenance_matricule = clean_text(row[idx['Institution origine']])
        niveau_etudes = clean_text(row[idx['Niveau etudes']])
        domaine_etudes = clean_text(row[idx['Domaine etudes']])
        sexe = (clean_text(row[idx['Sexe']]) or '').upper()[:1] or None
        statut = normalize_status(row[idx['Statut']])
        lieu_naissance = clean_text(row[idx['Lieu naissance']]) or 'Non renseigne'
        nombre_enfants = 0
        enfants_value = clean_text(row[idx['Enfants']])
        if enfants_value and enfants_value.isdigit():
            nombre_enfants = int(enfants_value)

        if target_id is not None:
            matricule_pnmls = fetch_existing_pnmls(cur, target_id)
        else:
            matricule_pnmls = next_pnmls_code(row_number)

        payload = {
            'matricule_pnmls': matricule_pnmls,
            'matricule_etat': matricule_etat,
            'provenance_matricule': provenance_matricule,
            'niveau_etudes': niveau_etudes,
            'annee_engagement_programme': engagement_year,
            'annee_naissance': birth_year,
            'nom': nom,
            'postnom': postnom,
            'organe': organe,
            'prenom': prenom,
            'email': primary_email,
            'email_prive': email_prive,
            'email_professionnel': email_professionnel,
            'password': None,
            'photo': None,
            'date_naissance': birth_date.isoformat(),
            'lieu_naissance': lieu_naissance,
            'situation_familiale': 'célibataire',
            'nombre_enfants': nombre_enfants,
            'telephone': telephone,
            'adresse': None,
            'poste_actuel': poste_actuel,
            'fonction': fonction,
            'grade_etat': grade_etat,
            'sexe': sexe,
            'departement_id': None,
            'province_id': province_id,
            'role_id': None,
            'grade_id': grade_id,
            'date_embauche': hire_date.isoformat(),
            'statut': statut,
        }

        if target_id is None:
            action_report['to_insert'] += 1
            if not args.dry_run:
                columns = ', '.join(payload.keys()) + ', created_at, updated_at'
                placeholders = ', '.join(['%s'] * len(payload)) + ', NOW(), NOW()'
                cur.execute(
                    f'INSERT INTO agents ({columns}) VALUES ({placeholders})',
                    list(payload.values()),
                )
                new_id = cur.lastrowid
                if matricule_etat:
                    existing_by_matricule[normalize_text(matricule_etat)] = new_id
                existing_by_email[primary_email] = new_id
                existing_by_identity[(normalize_text(nom), normalize_text(postnom), normalize_text(prenom), province_id or 0, normalize_text(organe))] = (new_id, matricule_pnmls)
                action_report['inserted'] += 1
        else:
            action_report['to_update'] += 1
            if not args.dry_run:
                assignments = ', '.join(f'{column} = %s' for column in payload.keys()) + ', updated_at = NOW()'
                cur.execute(
                    f'UPDATE agents SET {assignments} WHERE id = %s',
                    list(payload.values()) + [target_id],
                )
                existing_by_email[primary_email] = target_id
                action_report['updated'] += 1

    if args.dry_run:
        conn.rollback()
    else:
        conn.commit()

    report = {
        'file': str(workbook_path),
        'dry_run': args.dry_run,
        'db_host': args.db_host,
        'db_port': args.db_port,
        'db_name': args.db_name,
        'created_provinces': created_provinces,
        'action_report': dict(action_report),
        'anomaly_report': dict(anomaly_report),
        'total_rows': len(rows),
    }

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()